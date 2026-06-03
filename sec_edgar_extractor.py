"""
SEC EDGAR 10-K Financial Data Extractor
Pulls income statement and balance sheet data via XBRL API and exports to Excel.
"""

import sys
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict


# ── EDGAR API ──────────────────────────────────────────────────────────────────

HEADERS = {"User-Agent": "FinancialAnalysis tool@example.com"}


def get_cik(ticker: str) -> str:
    url = "https://www.sec.gov/files/company_tickers.json"
    resp = requests.get(url, headers=HEADERS)
    resp.raise_for_status()
    for entry in resp.json().values():
        if entry["ticker"].upper() == ticker.upper():
            return str(entry["cik_str"]).zfill(10)
    raise ValueError(f"Ticker '{ticker}' not found in EDGAR.")


def get_company_info(cik: str) -> tuple:
    """Return (company_name, sic_code) from the EDGAR submissions endpoint."""
    url = f"https://data.sec.gov/submissions/CIK{cik}.json"
    resp = requests.get(url, headers=HEADERS)
    resp.raise_for_status()
    data = resp.json()
    name = data.get("name", "Unknown Company")
    sic  = int(data.get("sic", 0) or 0)
    return name, sic


# SIC codes for oil, gas, and mining exploration/production companies
# 1311 Crude Petroleum & Natural Gas, 1381-1389 Oil & Gas Field Services,
# 1311-1382 range covers upstream E&P broadly
OIL_GAS_SIC_CODES = {
    1311,  # Crude Petroleum & Natural Gas
    1381,  # Drilling Oil & Gas Wells
    1382,  # Oil & Gas Field Services
    1389,  # Services-Oil & Gas Field, NEC
    2911,  # Petroleum Refining
    5171,  # Petroleum & Petroleum Products Wholesalers
    5172,  # Petroleum & Petroleum Products Wholesalers, NEC
    1321,  # Natural Gas Liquids
}


def ebitda_label(sic: int) -> str:
    """Return 'EBITDAX' for oil & gas SIC codes, 'EBITDA' for all others."""
    return "EBITDAX" if sic in OIL_GAS_SIC_CODES else "EBITDA"


def get_facts(cik: str) -> dict:
    url = f"https://data.sec.gov/api/xbrl/companyfacts/CIK{cik}.json"
    resp = requests.get(url, headers=HEADERS)
    resp.raise_for_status()
    return resp.json().get("facts", {})


def extract_annual_values(facts: dict, concept: str, years: list) -> dict:
    """Return {year: value} for a given XBRL concept from 10-K annual filings."""
    result = {y: None for y in years}
    namespaces = ["us-gaap", "ifrs-full", "dei"]
    for ns in namespaces:
        units = facts.get(ns, {}).get(concept, {}).get("units", {})
        for unit_key in ("USD", "shares"):
            entries = units.get(unit_key, [])
            # Keep only 10-K annual entries.
            # Balance sheet items have start=None (point-in-time); income items
            # have start set and must span ~12 months to exclude quarterly entries.
            annual = [
                e for e in entries
                if e.get("form") in ("10-K", "10-K/A") and e.get("end")
                and (
                    e.get("start") is None
                    or _months_between(e["start"], e["end"]) in (11, 12, 13)
                )
            ]
            # For each target year pick the most recently filed value
            by_year: dict = defaultdict(list)
            for e in annual:
                y = int(e["end"][:4])
                if y in years:
                    by_year[y].append(e)
            for y, candidates in by_year.items():
                if result[y] is None and candidates:
                    best = max(candidates, key=lambda e: e.get("filed", ""))
                    result[y] = best["val"]
            if any(v is not None for v in result.values()):
                return result
    return result


def _months_between(start: str, end: str) -> int:
    sy, sm = int(start[:4]), int(start[5:7])
    ey, em = int(end[:4]), int(end[5:7])
    return (ey - sy) * 12 + (em - sm)


# ── DATA CONCEPTS ──────────────────────────────────────────────────────────────

# Each concept list is tried in order; first non-None value wins.
INCOME_CONCEPTS = {
    "Revenue": [
        "Revenues",
        "RevenueFromContractWithCustomerExcludingAssessedTax",
        "SalesRevenueNet",
        "RevenueFromContractWithCustomerIncludingAssessedTax",
        "OilAndGasRevenue",
    ],
    "Total Expenses": [
        "OperatingExpenses",
        "CostsAndExpenses",
        "BenefitsLossesAndExpenses",
    ],
    "Depreciation / Depletion / Amortization": [
        "DepletionDepreciationAndAmortization",
        "DepreciationDepletionAndAmortization",
        "DepreciationAndAmortization",
        "Depreciation",
    ],
    "Other Income / Expense": [
        "NonoperatingIncomeExpense",
        "OtherNonoperatingIncomeExpense",
        "OtherOperatingIncomeExpense",
    ],
    "Net Income": [
        "NetIncomeLoss",
        "NetIncomeLossAvailableToCommonStockholdersBasic",
        "ProfitLoss",
    ],
    "Distributions": [
        "PaymentsOfDividends",
        "PaymentsOfDividendsCommonStock",
        "DistributionMadeToLimitedPartnerCashDistributionsPaid",
    ],
}

BALANCE_CONCEPTS = {
    "Cash and Cash Equivalents": [
        "CashAndCashEquivalentsAtCarryingValue",
        "Cash",
        "CashCashEquivalentsAndShortTermInvestments",
    ],
    "Other Current Assets": [
        "OtherAssetsCurrent",
        "PrepaidExpenseAndOtherAssetsCurrent",
    ],
    "Total Current Assets": ["AssetsCurrent"],
    "Fixed Assets, Net": [
        "PropertyPlantAndEquipmentNet",
        "PropertyPlantAndEquipmentAndFinanceLeaseRightOfUseAssetAfterAccumulatedDepreciationAndAmortization",
    ],
    "Other Non-Current Assets": [
        "OtherAssetsNoncurrent",
        "IntangibleAssetsNetExcludingGoodwill",
    ],
    "Total Assets": ["Assets"],
    "Notes Payable": [
        "NotesPayableCurrent",
        "ShortTermBorrowings",
        "CommercialPaper",
    ],
    "Accounts Payable and Other Current Liabilities": [
        "AccountsPayableAndAccruedLiabilitiesCurrent",
        "AccountsPayableCurrent",
        "LiabilitiesCurrent",  # fallback
    ],
    "Total Current Liabilities": ["LiabilitiesCurrent"],
    "Long-Term Debt": [
        "LongTermDebtNoncurrent",
        "LongTermDebt",
        "LongTermDebtAndCapitalLeaseObligations",
    ],
    "Other Liabilities": [
        "OtherLiabilitiesNoncurrent",
        "DeferredRevenueNoncurrent",
        "OtherLiabilities",
    ],
    "Total Liabilities": ["Liabilities"],
    "Total Equity": [
        "StockholdersEquity",
        "StockholdersEquityIncludingPortionAttributableToNoncontrollingInterest",
        "PartnersCapital",
    ],
    "Total Liabilities and Equity": [
        "LiabilitiesAndStockholdersEquity",
        "LiabilitiesAndPartnersCapital",
    ],
}


def fetch_row(facts: dict, concept_list: list, years: list) -> dict:
    """Try each concept in order; merge results preferring earlier concepts."""
    merged = {y: None for y in years}
    for concept in concept_list:
        vals = extract_annual_values(facts, concept, years)
        for y in years:
            if merged[y] is None and vals[y] is not None:
                merged[y] = vals[y]
        if all(v is not None for v in merged.values()):
            break
    return merged


def compute_ebitdax(income_data: dict, years: list) -> dict:
    """EBITDAX = Net Income + D&A + Interest Expense + Tax - Other Income (approx)."""
    # Simplified: EBITDA = Net Income + D&A (we add back what we have)
    ni = income_data.get("Net Income", {})
    da = income_data.get("Depreciation / Depletion / Amortization", {})
    result = {}
    for y in years:
        n = ni.get(y)
        d = da.get(y)
        if n is not None and d is not None:
            result[y] = n + d
        elif n is not None:
            result[y] = n
        else:
            result[y] = None
    return result


# ── EXCEL FORMATTING ───────────────────────────────────────────────────────────

GRAY_FILL = PatternFill("solid", fgColor="D9D9D9")
LIGHT_GRAY = PatternFill("solid", fgColor="F2F2F2")
THIN = Side(style="thin")
BORDER_BOTTOM = Border(bottom=Side(style="medium"))
BORDER_THIN_BOTTOM = Border(bottom=THIN)

BOLD = Font(bold=True)
BOLD_SMALL = Font(bold=True, size=9)
NORMAL = Font(size=9)
ITALIC = Font(italic=True, size=9)
TITLE_FONT = Font(bold=True, size=11)
HEADER_FONT = Font(bold=True, size=9)


def fmt_num(v) -> str:
    if v is None:
        return "N/A"
    return f"{v:,.0f}"


def avg(values: list) -> object:
    valid = [v for v in values if v is not None]
    return sum(valid) / len(valid) if valid else None


def set_cell(ws, row, col, value, font=None, fill=None, align=None, border=None, number_format=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:
        c.font = font
    if fill:
        c.fill = fill
    if align:
        c.alignment = align
    if border:
        c.border = border
    if number_format:
        c.number_format = number_format
    return c


def write_header_block(ws, start_row: int, company_name: str, sheet_title: str, years: list):
    """Write title + column headers, return next available row."""
    # Title row
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=2 + len(years) * 2 + 2)
    set_cell(ws, start_row, 1,
             f"{sheet_title} - {company_name}",
             font=TITLE_FONT, fill=GRAY_FILL,
             align=Alignment(horizontal="center"))
    r = start_row + 1

    # Subtitle
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2 + len(years) * 2 + 2)
    set_cell(ws, r, 1, "Year Ended December 31",
             font=Font(bold=True, size=9, italic=True), fill=GRAY_FILL,
             align=Alignment(horizontal="center"))
    r += 1

    # unit note row
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2 + len(years) * 2 + 2)
    set_cell(ws, r, 1, "(in thousands)",
             font=Font(italic=True, size=8), fill=GRAY_FILL,
             align=Alignment(horizontal="center"))
    r += 1

    # Column headers
    col = 1
    set_cell(ws, r, col, "", font=HEADER_FONT, fill=GRAY_FILL)
    col += 1
    set_cell(ws, r, col, "(in U.S. $)", font=HEADER_FONT, fill=GRAY_FILL)
    col += 1

    col_labels = [str(y) for y in years] + ["Average"]
    for label in col_labels:
        ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col + 1)
        set_cell(ws, r, col, label,
                 font=HEADER_FONT, fill=GRAY_FILL,
                 align=Alignment(horizontal="center"))
        col += 2

    r += 1

    # Sub-header: $ / %
    col = 3
    for _ in col_labels:
        set_cell(ws, r, col, "$", font=HEADER_FONT, fill=GRAY_FILL,
                 align=Alignment(horizontal="right"))
        set_cell(ws, r, col + 1, "%", font=HEADER_FONT, fill=GRAY_FILL,
                 align=Alignment(horizontal="right"))
        col += 2

    r += 1
    return r


def write_data_row(ws, row: int, label: str, values: list,
                   bold=False, shaded=False, italic=False, base_values: list = None):
    """
    values:       raw EDGAR dollar values (full dollars) for each year.
    base_values:  raw EDGAR dollar values for the base row (Revenue / Total Assets).
                  Used to compute common-size percentages.

    All values are written as static numbers so every application (Excel,
    Numbers, LibreOffice) displays them correctly without formula recalculation.
    Percentages are stored as decimals (e.g. 0.253) with number_format '0.0%'.
    """
    fill = GRAY_FILL if shaded else None
    font = BOLD if bold else (ITALIC if italic else NORMAL)
    header_font = BOLD_SMALL if bold else Font(size=9, italic=italic)

    set_cell(ws, row, 1, "", fill=fill)
    set_cell(ws, row, 2, label, font=header_font, fill=fill,
             align=Alignment(indent=0 if bold else 1))

    n_years = len(values)
    dollar_cols = [3 + i * 2 for i in range(n_years)]
    avg_dollar_col = 3 + n_years * 2

    # ── Year dollar values ────────────────────────────────────────────────
    for i, v in enumerate(values):
        col = dollar_cols[i]
        if v is not None:
            set_cell(ws, row, col, v / 1_000_000,
                     font=font, fill=fill,
                     align=Alignment(horizontal="right"),
                     number_format='#,##0')
        else:
            set_cell(ws, row, col, None, font=font, fill=fill,
                     align=Alignment(horizontal="right"))

    # ── Average dollar value (Python-computed) ────────────────────────────
    valid = [v for v in values if v is not None]
    avg_val = (sum(valid) / len(valid) / 1_000_000) if valid else None
    set_cell(ws, row, avg_dollar_col,
             avg_val if avg_val is not None else None,
             font=font, fill=fill,
             align=Alignment(horizontal="right"),
             number_format='#,##0' if avg_val is not None else None)

    # ── Common-size % columns (Python-computed decimals) ──────────────────
    if base_values is not None:
        valid_bases = [b for b in base_values if b is not None]
        avg_base = sum(valid_bases) / len(valid_bases) if valid_bases else None

        for i, (v, b) in enumerate(zip(values, base_values)):
            pct = (v / b) if (v is not None and b) else None
            col = dollar_cols[i] + 1
            set_cell(ws, row, col,
                     pct if pct is not None else None,
                     font=font, fill=fill,
                     align=Alignment(horizontal="right"),
                     number_format='0.0%' if pct is not None else None)

        # Average %
        avg_pct = (avg_val * 1_000_000 / avg_base) if (avg_val is not None and avg_base) else None
        set_cell(ws, row, avg_dollar_col + 1,
                 avg_pct if avg_pct is not None else None,
                 font=font, fill=fill,
                 align=Alignment(horizontal="right"),
                 number_format='0.0%' if avg_pct is not None else None)


def write_footer(ws, row: int, num_cols: int):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    set_cell(ws, row, 1, "Source: Compiled Financial Statements",
             font=ITALIC, align=Alignment(horizontal="left"))


def set_column_widths(ws, years_count: int):
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 38
    col = 3
    for _ in range(years_count + 1):  # years + avg
        ws.column_dimensions[get_column_letter(col)].width = 14
        ws.column_dimensions[get_column_letter(col + 1)].width = 8
        col += 2


# ── SHEET WRITERS ──────────────────────────────────────────────────────────────

def write_income_sheet(wb, company_name: str, years: list, income_data: dict,
                       ebitda_row_label: str = "EBITDA"):
    ws = wb.active
    ws.title = "Income Statement"
    num_cols = 2 + (len(years) + 1) * 2

    r = write_header_block(ws, 1, company_name,
                           "Income Statements ($) and Common Size-Equivalents", years)

    revenue_vals = [income_data.get("Revenue", {}).get(y) for y in years]

    rows = [
        ("Income / Revenue",                          "Revenue",                              True),
        ("Operating Expenses",                        "Total Expenses",                       True),
        (ebitda_row_label,                            "EBITDAX",                              True),
        ("Depreciation / Depletion / Amortization",  "Depreciation / Depletion / Amortization", False),
        ("Other Income / Expense",                   "Other Income / Expense",               False),
        ("Net Income",                               "Net Income",                           True),
        ("Distributions",                            "Distributions",                        False),
    ]

    for label, key, bold in rows:
        vals = [income_data.get(key, {}).get(y) for y in years]
        write_data_row(ws, r, label, vals, bold=bold, base_values=revenue_vals)
        r += 1

    write_footer(ws, r + 1, num_cols)
    set_column_widths(ws, len(years))


def write_balance_sheet(wb, company_name: str, years: list, bs_data: dict):
    ws = wb.create_sheet("Balance Sheet")
    num_cols = 2 + (len(years) + 1) * 2

    r = write_header_block(ws, 1, company_name,
                           "Balance Sheets ($) and Common Size-Equivalents", years)

    total_assets_vals = [bs_data.get("Total Assets", {}).get(y) for y in years]

    all_rows = [
        # (display label,               data key,                                   bold,  section_header)
        ("Assets",                       None,                                        True,  True),
        ("Cash and Cash Equivalents",    "Cash and Cash Equivalents",                 False, False),
        ("Other Current Assets",         "Other Current Assets",                      False, False),
        ("Total Current Assets",         "Total Current Assets",                      True,  False),
        ("Fixed Assets, Net",            "Fixed Assets, Net",                         False, False),
        ("Other Non-Current Assets",     "Other Non-Current Assets",                  False, False),
        ("Total Assets",                 "Total Assets",                              True,  False),
        ("Liabilities and Equity",       None,                                        True,  True),
        ("Notes Payable",                "Notes Payable",                             False, False),
        ("Accounts Payable and Other Current Liabilities",
                                         "Accounts Payable and Other Current Liabilities", False, False),
        ("Total Current Liabilities",    "Total Current Liabilities",                 True,  False),
        ("Long-Term Debt",               "Long-Term Debt",                            False, False),
        ("Other Liabilities",            "Other Liabilities",                         False, False),
        ("Total Liabilities",            "Total Liabilities",                         True,  False),
        ("Total Equity",                 "Total Equity",                              True,  False),
        ("Total Liabilities and Equity", "Total Liabilities and Equity",              True,  False),
    ]

    for label, key, bold, section_header in all_rows:
        if section_header:
            set_cell(ws, r, 1, "", fill=GRAY_FILL)
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=num_cols)
            set_cell(ws, r, 2, label, font=BOLD, fill=GRAY_FILL)
            r += 1
            continue
        vals = [bs_data.get(key, {}).get(y) for y in years]
        write_data_row(ws, r, label, vals, bold=bold, base_values=total_assets_vals)
        r += 1

    write_footer(ws, r + 1, num_cols)
    set_column_widths(ws, len(years))


# ── MAIN ───────────────────────────────────────────────────────────────────────

def build_workbook(ticker: str, num_years: int = 3) -> str:
    print(f"Looking up CIK for {ticker.upper()}...")
    cik = get_cik(ticker)
    company_name, sic = get_company_info(cik)
    label = ebitda_label(sic)
    print(f"Found: {company_name} (CIK {cik}, SIC {sic}) → using '{label}'")

    print("Fetching XBRL facts...")
    facts = get_facts(cik)

    # Determine the most recent 3 fiscal years available for Net Income
    print("Resolving fiscal years...")
    probe = fetch_row(facts, INCOME_CONCEPTS["Net Income"], list(range(2010, 2026)))
    available_years = sorted([y for y, v in probe.items() if v is not None], reverse=True)
    if not available_years:
        raise ValueError("No annual Net Income data found for this company.")
    years = sorted(available_years[:num_years])
    print(f"Using years: {years}")

    # Fetch income statement
    print("Fetching income statement data...")
    income_data: dict = {}
    for row_label, concepts in INCOME_CONCEPTS.items():
        income_data[row_label] = fetch_row(facts, concepts, years)
    income_data["EBITDAX"] = compute_ebitdax(income_data, years)

    # Fetch balance sheet
    print("Fetching balance sheet data...")
    bs_data: dict = {}
    for row_label, concepts in BALANCE_CONCEPTS.items():
        bs_data[row_label] = fetch_row(facts, concepts, years)

    # Build Excel
    print("Building Excel workbook...")
    wb = openpyxl.Workbook()
    write_income_sheet(wb, company_name, years, income_data, ebitda_row_label=label)
    write_balance_sheet(wb, company_name, years, bs_data)

    filename = f"{ticker.upper()}_financials.xlsx"
    wb.save(filename)
    print(f"\nSaved: {filename}")
    return filename


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python sec_edgar_extractor.py <TICKER> [num_years]")
        print("  e.g. python sec_edgar_extractor.py AAPL")
        print("       python sec_edgar_extractor.py XOM 5")
        sys.exit(1)

    ticker_arg = sys.argv[1]
    years_arg = int(sys.argv[2]) if len(sys.argv) > 2 else 3
    build_workbook(ticker_arg, years_arg)

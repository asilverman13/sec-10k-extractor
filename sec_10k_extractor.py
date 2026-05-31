#!/usr/bin/env python3
"""
SEC EDGAR 10-K Financial Data Extractor

Usage:
    python3 sec_10k_extractor.py AAPL
    python3 sec_10k_extractor.py AAPL,MSFT,GOOG
    python3 sec_10k_extractor.py XOM,CVX --years 3 --export results.xlsx
"""

from __future__ import annotations

import re
import sys
import time
import argparse
import requests
from datetime import datetime
from typing import Optional
from bs4 import BeautifulSoup
from tabulate import tabulate

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

HEADERS = {
    "User-Agent": "FinancialResearchTool adogggoat13@icloud.com",
    "Accept-Encoding": "gzip, deflate",
    "Host": "data.sec.gov",
}
HTML_HEADERS = {
    "User-Agent": "FinancialResearchTool adogggoat13@icloud.com",
    "Accept-Encoding": "gzip, deflate",
}

OIL_GAS_SIC = {
    "1311", "1321", "1381", "1382", "1389", "2911", "5171", "5172",
}

# Metrics displayed in order; oil/gas-only ones are filtered later
METRIC_KEYS = [
    "revenue", "ebitda", "net_income", "total_assets",
    "total_debt", "cash", "equity", "pv10", "proved_reserves",
]
METRIC_LABELS = {
    "revenue":         "Revenue",
    "ebitda":          "EBITDA (est.)",
    "net_income":      "Net Income",
    "total_assets":    "Total Assets",
    "total_debt":      "Total Debt",
    "cash":            "Cash",
    "equity":          "Equity",
    "pv10":            "PV-10",
    "proved_reserves": "Proved Reserves",
}

TRUST_METRIC_LABELS = {
    "revenue":         "Royalty Income",
    "ebitda":          "EBITDA (est.)",
    "net_income":      "Distributable Income",
    "total_assets":    "Total Assets",
    "total_debt":      "Total Debt",
    "cash":            "Cash",
    "equity":          "Trust Corpus",
    "pv10":            "PV-10 (Std. Measure)",
    "proved_reserves": "Proved Reserves",
}

# ---------------------------------------------------------------------------
# EDGAR lookup helpers
# ---------------------------------------------------------------------------

# Known tickers that EDGAR doesn't register but whose CIKs are stable
_KNOWN_CIKS: dict[str, tuple[str, str]] = {
    "BPT":   ("0000850033", "BP PRUDHOE BAY ROYALTY TRUST"),
    "SDT":   ("0001521168", "SANDRIDGE PERMIAN TRUST"),
    "TIRTZ": ("0001581552", "TORCHLIGHT ENERGY ROYALTY TRUST"),
}


def get_cik(ticker: str) -> tuple[str, str]:
    """Return (cik_padded, company_name) for a ticker."""
    ticker_upper = ticker.upper()

    # 1. Manual override for tickers EDGAR doesn't index
    if ticker_upper in _KNOWN_CIKS:
        cik, name = _KNOWN_CIKS[ticker_upper]
        return cik.zfill(10), name

    # 2. Standard company_tickers.json
    url = "https://www.sec.gov/files/company_tickers.json"
    r = requests.get(url, headers=HTML_HEADERS, timeout=15)
    r.raise_for_status()
    for entry in r.json().values():
        if entry["ticker"].upper() == ticker_upper:
            return str(entry["cik_str"]).zfill(10), entry["title"]

    # 3. Broader exchange file (includes OTC / pink-sheet tickers)
    r2 = requests.get("https://www.sec.gov/files/company_tickers_exchange.json",
                      headers=HTML_HEADERS, timeout=15)
    if r2.ok:
        for row in r2.json().get("data", []):
            # row: [cik, name, ticker, exchange]
            if len(row) >= 3 and str(row[2]).upper() == ticker_upper:
                return str(row[0]).zfill(10), str(row[1])

    # 4. Full-text EDGAR search by ticker symbol
    r3 = requests.get(
        f"https://efts.sec.gov/LATEST/search-index?q=%22{ticker_upper}%22&forms=10-K",
        headers=HTML_HEADERS, timeout=10,
    )
    if r3.ok:
        hits = r3.json().get("hits", {}).get("hits", [])
        for h in hits:
            src = h.get("_source", {})
            entity_id = src.get("entity_id") or src.get("cik")
            entity_name = src.get("entity_name", "")
            if entity_id:
                return str(entity_id).zfill(10), entity_name

    raise ValueError(f"Ticker '{ticker}' not found in SEC EDGAR.")


def get_company_info(cik: str) -> dict:
    url = f"https://data.sec.gov/submissions/CIK{cik}.json"
    r = requests.get(url, headers=HEADERS, timeout=15)
    r.raise_for_status()
    return r.json()


def get_10k_filings(submissions: dict, n: int = 3) -> list[dict]:
    """
    Return the most recent `n` distinct fiscal-year 10-K filings as a list of
    {"accession": ..., "date": ..., "original_accession": ..., "fiscal_year": ...}

    When a 10-K/A is the newest filing for a fiscal year, we also record the
    original 10-K accession so we can fall back to it for oil-and-gas HTML.
    """
    filings = submissions.get("filings", {}).get("recent", {})
    forms   = filings.get("form", [])
    accs    = filings.get("accessionNumber", [])
    dates   = filings.get("filingDate", [])
    periods = filings.get("reportDate", dates)   # fiscal year-end date

    # Group by fiscal year-end (reportDate), keeping only 10-K / 10-K/A
    by_year: dict[str, dict] = {}
    for form, acc, date, period in zip(forms, accs, dates, periods):
        if form not in ("10-K", "10-K/A"):
            continue
        fiscal_yr = period[:4] if period else date[:4]
        if fiscal_yr not in by_year:
            by_year[fiscal_yr] = {"accession": acc, "date": date,
                                   "original_accession": None, "fiscal_year": fiscal_yr}
        # Track the first plain 10-K so we can fall back for HTML scraping
        if form == "10-K" and by_year[fiscal_yr]["original_accession"] is None:
            by_year[fiscal_yr]["original_accession"] = acc

    # Sort descending by fiscal year, take most recent n
    years = sorted(by_year.keys(), reverse=True)[:n]
    result = [by_year[y] for y in years]

    # Fill original_accession with accession itself if no amendment existed
    for r in result:
        if r["original_accession"] is None:
            r["original_accession"] = r["accession"]

    if not result:
        raise ValueError("No 10-K filings found for this company.")
    return result


# ---------------------------------------------------------------------------
# XBRL helpers
# ---------------------------------------------------------------------------

def get_company_facts(cik: str) -> dict:
    url = f"https://data.sec.gov/api/xbrl/companyfacts/CIK{cik}.json"
    r = requests.get(url, headers=HEADERS, timeout=20)
    if r.status_code == 404:
        return {}
    r.raise_for_status()
    return r.json()


def _annual_entries(facts: dict, concept: str, namespace: str = "us-gaap") -> list[dict]:
    """All annual 10-K/10-K/A entries for a concept, sorted newest-first."""
    try:
        units = facts["facts"][namespace][concept]["units"]
        entries = units.get("USD", units.get("pure", []))
        annual = [e for e in entries
                  if e.get("form") in ("10-K", "10-K/A") and e.get("val") is not None]
        annual.sort(key=lambda e: e["end"], reverse=True)
        return annual
    except (KeyError, TypeError):
        return []


def value_for_year(facts: dict, concepts: list[str], fiscal_year: str,
                   namespace: str = "us-gaap") -> Optional[float]:
    """
    Return the value for one of the given XBRL concepts whose fiscal-year-end
    matches `fiscal_year` (a 4-digit string like "2023").
    Returns None if no concept has data for that exact year.
    """
    for concept in concepts:
        for e in _annual_entries(facts, concept, namespace):
            if e["end"].startswith(fiscal_year):
                return e["val"]
    return None


def _build_series(facts: dict, concepts: list[str]) -> dict[str, Optional[float]]:
    """
    Return {fiscal_year_str: value} covering all years present in XBRL.
    First concept that has data for a year wins.
    """
    series: dict[str, Optional[float]] = {}
    for concept in concepts:
        for e in _annual_entries(facts, concept):
            yr = e["end"][:4]
            if yr not in series:
                series[yr] = e["val"]
    return series


def _build_ebitda_series(facts: dict) -> dict[str, Optional[float]]:
    """
    EBITDA = Operating/pre-tax income + D&A, matched by fiscal year.
    """
    income_concepts = [
        "OperatingIncomeLoss",
        "IncomeLossFromContinuingOperationsBeforeIncomeTaxesExtraordinaryItemsNoncontrollingInterest",
        "IncomeLossFromContinuingOperationsBeforeIncomeTaxesMinorityInterestAndIncomeLossFromEquityMethodInvestments",
    ]
    da_concepts = [
        "DepreciationDepletionAndAmortization",
        "DepreciationAndAmortization",
        "Depreciation",
    ]
    income_series = _build_series(facts, income_concepts)
    da_series     = _build_series(facts, da_concepts)
    common = set(income_series) & set(da_series)
    return {yr: income_series[yr] + da_series[yr] for yr in common}  # type: ignore[operator]


# ---------------------------------------------------------------------------
# Unit-scale detection & normalisation
# ---------------------------------------------------------------------------

def detect_xbrl_scale(facts: dict) -> int:
    """
    Determine the unit multiplier from XBRL data.

    EDGAR XBRL values are always in *base units* (dollars, not thousands/millions).
    So the scale is always 1 — but we keep this function to document that assumption
    and to handle any future edge case.
    """
    return 1  # XBRL values from data.sec.gov are always in whole dollars


def detect_html_scale(html_text: str) -> int:
    """
    Return the multiplier to convert the filing's reported numbers to whole dollars.
    Scans for the disclosure phrase that governs dollar amounts (not shares).
    """
    low = html_text.lower()
    # Match "in millions" referring to dollar values (most large-cap filings)
    if re.search(r"(?:stated|expressed|reported|presented|denominated|"
                 r"amounts? are stated|values?.{0,30})\s*in millions", low):
        return 1_000_000
    # "in thousands" for dollar amounts (exclude "(shares in thousands)" etc.)
    if re.search(r"(?:stated|expressed|reported|presented|denominated|"
                 r"amounts? are stated|values?.{0,30})\s*in thousands", low):
        return 1_000
    # Bare "in thousands" / "in millions" — accept only if share context is absent nearby
    for m in re.finditer(r"\bin thousands\b", low):
        ctx = low[max(0, m.start() - 60): m.start() + 40]
        if "share" not in ctx and "stock" not in ctx and "unit" not in ctx:
            return 1_000
    for m in re.finditer(r"\bin millions\b", low):
        ctx = low[max(0, m.start() - 60): m.start() + 40]
        if "share" not in ctx and "stock" not in ctx and "unit" not in ctx:
            return 1_000_000
    return 1_000_000   # safe default for large-cap filings


def to_thousands(value: Optional[float]) -> Optional[float]:
    """Convert a whole-dollar XBRL value to thousands (our standard display unit)."""
    if value is None:
        return None
    return value / 1_000


# ---------------------------------------------------------------------------
# HTML filing fetch
# ---------------------------------------------------------------------------

def _resolve_doc_url(href: str) -> str:
    if href.startswith("/ix?doc="):
        href = href[len("/ix?doc="):]
    return href if href.startswith("http") else f"https://www.sec.gov{href}"


def fetch_10k_html(cik: str, accession: str) -> str:
    acc_nodash = accession.replace("-", "")
    index_url = (f"https://www.sec.gov/Archives/edgar/data/{int(cik)}"
                 f"/{acc_nodash}/{accession}-index.htm")
    r = requests.get(index_url, headers=HTML_HEADERS, timeout=15)
    soup = BeautifulSoup(r.text, "lxml")

    primary = None
    for row in soup.select("table tr"):
        cells = row.find_all("td")
        if len(cells) >= 4 and "10-K" in cells[3].get_text():
            link = cells[2].find("a")
            if link:
                href = link["href"]
                if "ix?doc=" in href or href.endswith((".htm", ".html")):
                    primary = _resolve_doc_url(href)
                    break
    if not primary:
        for a in soup.select("table a[href]"):
            href = a["href"]
            if "ix?doc=" in href or href.endswith((".htm", ".html")):
                primary = _resolve_doc_url(href)
                break
    if not primary:
        return ""

    time.sleep(0.3)
    r2 = requests.get(primary, headers=HTML_HEADERS, timeout=60)
    return r2.text


# ---------------------------------------------------------------------------
# HTML table scraping helpers (for PV-10 and reserves)
# ---------------------------------------------------------------------------

def _parse_num(s: str) -> Optional[float]:
    s = s.strip().replace(",", "").replace("$", "")
    negative = s.startswith("(") and s.endswith(")")
    s = s.strip("()")
    try:
        val = float(s)
        return -val if negative else val
    except ValueError:
        return None


def scrape_table_value(soup: BeautifulSoup, label_patterns: list[str]) -> Optional[float]:
    """Search all table cells for a label match, return the first numeric sibling."""
    for tag in soup.find_all(["td", "th"]):
        cell_text = tag.get_text(" ", strip=True).lower()
        if not any(re.search(p, cell_text) for p in label_patterns):
            continue
        row = tag.find_parent("tr")
        if not row:
            continue
        cells = list(row.find_all(["td", "th"]))
        try:
            idx = cells.index(tag)
        except ValueError:
            continue
        for td in cells[idx + 1:]:
            v = _parse_num(td.get_text(" ", strip=True))
            if v is not None and v != 0:
                return v
    return None


def scrape_number_from_text(text: str, patterns: list[str], window: int = 400) -> Optional[float]:
    low = text.lower()
    for pat in patterns:
        m = re.search(pat, low)
        if m:
            snippet = low[m.start(): m.start() + window]
            num_m = re.search(r"\(?([\d,]{3,}(?:\.\d+)?)\)?", snippet[len(m.group(0)):])
            if num_m:
                try:
                    val = float(num_m.group(1).replace(",", ""))
                    return -val if num_m.group(0).startswith("(") else val
                except ValueError:
                    continue
    return None


# ---------------------------------------------------------------------------
# Oil & gas HTML extraction (PV-10 + proved reserves)
# ---------------------------------------------------------------------------

def _reserves_year_end(plain_text: str, year: int) -> Optional[float]:
    """
    Find total proved reserves for `year` from the rollforward table.
    The last substantive number after 'December 31, {year}' is the BOE total column.
    """
    lower = plain_text.lower()
    start = lower.find("proved developed and undeveloped reserves")
    if start < 0:
        return None
    section = plain_text[start: start + 6_000]
    m = re.search(rf"december 31, {year}", section.lower())
    if not m:
        return None
    row_text = section[m.start(): m.start() + 300]
    nums = []
    for n in re.findall(r"[\d,]+(?:\.\d+)?", row_text):
        try:
            v = float(n.replace(",", ""))
            if v != year and v > 10:
                nums.append(v)
        except ValueError:
            pass
    return nums[-1] if nums else None


def scrape_pv10_and_reserves(html_text: str) -> tuple[Optional[float], Optional[float]]:
    """
    Return (pv10_raw, proved_reserves_raw) in the filing's own reported units.
    The caller is responsible for applying the HTML unit scale.
    """
    soup = BeautifulSoup(html_text, "html.parser")
    plain = soup.get_text(" ")

    pv10 = scrape_table_value(soup, [
        r"standardized measure of discounted future net cash flows",
        r"standardized measure of discounted",
        r"pv[\s\-]?10",
    ])
    if pv10 is None:
        pv10 = scrape_number_from_text(plain, [
            r"standardized measure of discounted future net cash flows",
            r"pv[\s\-]?10",
        ])

    current_year = datetime.now().year
    reserves = None
    for yr in (current_year - 1, current_year - 2):
        reserves = _reserves_year_end(plain, yr)
        if reserves is not None:
            break
    if reserves is None:
        reserves = scrape_table_value(soup, [r"total proved reserves", r"^total proved$"])
    if reserves is None:
        reserves = scrape_number_from_text(plain, [r"total proved reserves"])

    return pv10, reserves


# ---------------------------------------------------------------------------
# Royalty trust detection & HTML scraping
# ---------------------------------------------------------------------------

# Trust SIC codes + name keywords
_TRUST_SIC = {"6792", "6726"}


def is_royalty_trust(company_name: str, sic: str, facts: dict) -> bool:
    """
    Detect royalty trusts, which file under modified cash basis with no/minimal XBRL.
    """
    name_lower = company_name.lower()
    if "royalty trust" in name_lower or "oil royalty" in name_lower:
        return True
    if sic in _TRUST_SIC:
        return True
    # No XBRL data at all — strong signal for legacy trust filers
    gaap = facts.get("facts", {}).get("us-gaap", {})
    tagged = sum(1 for v in gaap.values()
                 if any(e.get("form") in ("10-K", "10-K/A")
                        for e in v.get("units", {}).get("USD", [])))
    return tagged == 0


def _scrape_trust_row(soup: BeautifulSoup, label_patterns: list[str],
                      max_cols: int = 3, min_val: float = 0,
                      prefer_largest: bool = False) -> list[float]:
    """
    Find a table row matching any label pattern and return up to max_cols
    numeric values (left-to-right = most-recent year first).
    Handles trust tables where dollar signs occupy their own <td>.
    Skips rows where the largest value is <= min_val (e.g. TOC page numbers).
    If prefer_largest=True, scans ALL matching rows and returns the one with
    the largest max value (useful when 'total' appears as multiple subtotals).
    """
    best: list[float] = []
    for tag in soup.find_all(["td", "th"]):
        cell_text = tag.get_text(" ", strip=True).lower()
        if not any(re.search(p, cell_text) for p in label_patterns):
            continue
        row = tag.find_parent("tr")
        if not row:
            continue
        cells = list(row.find_all(["td", "th"]))
        try:
            start = cells.index(tag) + 1
        except ValueError:
            continue
        vals = []
        for td in cells[start:]:
            t = td.get_text(" ", strip=True)
            v = _parse_num(t)
            if v is not None and abs(v) > 0:
                vals.append(v)
            if len(vals) >= max_cols:
                break
        # Skip rows whose values are all <= min_val (page numbers / tiny totals)
        if vals and max(abs(v) for v in vals) <= min_val:
            continue
        if vals:
            if not prefer_largest:
                return vals
            if not best or max(abs(v) for v in vals) > max(abs(v) for v in best):
                best = vals
    return best


def _plain_multi(plain: str, label: str, max_vals: int = 3) -> list[float]:
    """
    Regex fallback: find label in plain text, collect the next `max_vals`
    numbers that follow dollar signs.  Works for trusts whose tables
    render as   label … $ nnn,nnn  $ nnn,nnn  …
    """
    idx = plain.lower().find(label.lower())
    if idx < 0:
        return []
    snippet = plain[idx: idx + 600]
    # find numbers preceded by $
    hits = re.findall(r'\$\s*([\d,]+(?:\.\d+)?)', snippet)
    results = []
    for h in hits:
        try:
            v = float(h.replace(",", ""))
            if v > 0:
                results.append(v)
        except ValueError:
            pass
        if len(results) >= max_vals:
            break
    # If no $ signs, fall back to any large numbers
    if not results:
        nums = re.findall(r'\b([\d,]{4,}(?:\.\d+)?)\b', snippet)
        for n in nums:
            try:
                v = float(n.replace(",", ""))
                if v > 1_000:
                    results.append(v)
            except ValueError:
                pass
            if len(results) >= max_vals:
                break
    return results


def scrape_trust_all_years(
    html_text: str,
    filing_fiscal_year: str,
    is_oil_gas: bool,
) -> dict[str, dict]:
    """
    Parse a trust 10-K and extract up to 3 years of data from its
    comparative financial statements.  Returns {fiscal_year: metrics_dict}.

    Trust-specific terminology mapped to our standard keys:
        revenue       ← Royalty Income / Trust Income
        net_income    ← Distributable Income
        equity        ← Trust Corpus (end of year)
        total_assets  ← Total Assets (or Royalty Properties + Cash)
        cash          ← Cash / Cash and Cash Equivalents
        total_debt    ← 0 (trusts carry no debt)
        ebitda        ← None (not applicable)
    """
    soup  = BeautifulSoup(html_text, "html.parser")
    plain = soup.get_text(" ")

    # Royalty trusts ALWAYS report in whole dollars — never apply a scale multiplier.
    # (Phrases like "expressed in millions of cubic feet" refer to reserves units,
    #  not the financial statement currency, and must not affect dollar scale.)
    scale = 1

    def pick(patterns: list[str], n: int = 3, min_val: float = 0,
             prefer_largest: bool = False) -> list[float]:
        vals = _scrape_trust_row(soup, patterns, max_cols=n, min_val=min_val,
                                 prefer_largest=prefer_largest)
        if not vals:
            for pat in patterns:
                vals = _plain_multi(plain, pat, max_vals=n)
                if vals:
                    break
        return vals

    # --- Income statement (3-year comparative) ---
    # Use min_val=1000 to skip TOC rows (page numbers).
    # Use ^royalty income so "Future Royalty income" in oil/gas supplementals is skipped.
    revenue_vals    = pick([r"^royalty income", r"^trust income", r"^total income\b"],
                           min_val=1000)
    if not revenue_vals:
        revenue_vals = pick([r"royalty income", r"trust income"], min_val=1000)
    net_income_vals = pick([r"^distributable income$", r"^distributable income\b"],
                           min_val=1000)
    if not net_income_vals:
        net_income_vals = pick([r"distributable income", r"net income",
                                r"income available"], min_val=1000)

    # --- Balance sheet (2-year comparative) ---
    cash_vals = pick([r"cash and cash equiv", r"^cash\b"], n=2)

    # Trust corpus: "trust corpus$" would match TOC entry "Statements of Changes in
    # Trust Corpus", so anchor with ^ and require min_val to skip page numbers.
    corpus_vals = pick([r"trust corpus, end of year", r"^trust corpus\b",
                        r"corpus, end of year"], n=2, min_val=1000)

    liabilities_vals = pick([r"^total liabilities\b", r"total liabilities"], n=2)
    royalty_prop     = pick([r"royalty properties", r"net overriding royalty",
                             r"net royalty interests"], n=2)

    # Total assets: try explicit label first (many trusts use bare "Total").
    # Avoid prefer_largest — income statement totals are larger than balance-sheet totals.
    # Instead, derive from balance-sheet components when the label is absent.
    assets_vals = pick([r"^total assets\b", r"total assets"], n=2)
    if not assets_vals and corpus_vals and liabilities_vals:
        # Most reliable: balance sheet identity (Assets = Equity + Liabilities)
        assets_vals = [c + l for c, l in zip(corpus_vals, liabilities_vals)]
    if not assets_vals:
        # Last resort: "Total" row — take the FIRST one with plausible size
        assets_vals = pick([r"^total\b"], n=2, min_val=100_000)
    if not assets_vals and royalty_prop and cash_vals:
        assets_vals = [r + c for r, c in zip(royalty_prop, cash_vals)]

    # Sanity check: total assets must be >= trust corpus.  If not, we grabbed
    # an intermediate subtotal (e.g. royalty income total); fall back to derivation.
    if (assets_vals and corpus_vals
            and assets_vals[0] < corpus_vals[0] * 0.90):
        if liabilities_vals:
            assets_vals = [c + l for c, l in zip(corpus_vals, liabilities_vals)]
        else:
            assets_vals = []

    # --- Oil & gas supplemental (most recent year only from this filing) ---
    pv10_raw, reserves_raw = (None, None)
    if is_oil_gas:
        pv10_html, res_html = scrape_pv10_and_reserves(html_text)
        if pv10_html is not None:
            pv10_raw = pv10_html * scale
        reserves_raw = res_html

    # --- Map columns → fiscal years ---
    base_year = int(filing_fiscal_year)
    result: dict[str, dict] = {}

    def _safe(lst: list[float], i: int) -> Optional[float]:
        return lst[i] if i < len(lst) else None

    for offset in range(3):
        yr = str(base_year - offset)
        revenue    = _safe(revenue_vals, offset)
        net_income = _safe(net_income_vals, offset)
        cash       = _safe(cash_vals, offset)
        assets     = _safe(assets_vals, offset)
        corpus     = _safe(corpus_vals, offset)

        # Apply scale → whole dollars → convert to $thousands
        def ths(v: Optional[float]) -> Optional[float]:
            return to_thousands(v * scale) if v is not None else None

        result[yr] = {
            "revenue":         ths(revenue),
            "ebitda":          None,         # not applicable for trusts
            "net_income":      ths(net_income),
            "total_assets":    ths(assets),
            "total_debt":      0.0,          # trusts have no debt by design
            "cash":            ths(cash),
            "equity":          ths(corpus),
            "pv10":            to_thousands(pv10_raw) if offset == 0 else None,
            "proved_reserves": reserves_raw   if offset == 0 else None,
        }

    return result


# ---------------------------------------------------------------------------
# Per-year data extraction
# ---------------------------------------------------------------------------

def extract_year(
    facts: dict,
    fiscal_year: str,
    ebitda_series: dict[str, Optional[float]],
    is_oil_gas: bool,
    cik: str,
    accession: str,
    original_accession: str,
    scrape_html: bool = False,   # caller controls whether to fetch HTML for this year
) -> dict[str, Optional[float]]:
    """
    Pull all metrics for one fiscal year. All dollar values returned in $thousands.
    proved_reserves is returned as a raw float in the filing's own units (MMBoe/Bcfe).
    """

    def v(concepts: list[str]) -> Optional[float]:
        raw = value_for_year(facts, concepts, fiscal_year)
        return to_thousands(raw)

    revenue = v([
        "RevenueFromContractWithCustomerExcludingAssessedTax",
        "Revenues",
        "SalesRevenueNet",
        "RevenueFromContractWithCustomerIncludingAssessedTax",
        "SalesRevenueGoodsNet",
    ])
    net_income = v([
        "NetIncomeLoss",
        "NetIncomeLossAvailableToCommonStockholdersBasic",
        "ProfitLoss",
    ])
    total_assets = v(["Assets"])
    cash = v([
        "CashAndCashEquivalentsAtCarryingValue",
        "CashCashEquivalentsAndShortTermInvestments",
        "Cash",
    ])

    # Debt: try an explicit total first, then build from components
    total_debt_raw = value_for_year(facts, [
        "LongTermDebtAndCapitalLeaseObligations",   # XOM and others use this as full total
        "DebtAndCapitalLeaseObligations",
        "LongTermDebt",                             # some filers include current portion here
    ], fiscal_year)

    if total_debt_raw is not None:
        # May already include current; check if we should add separate current portion
        current_raw = value_for_year(facts, ["DebtCurrent", "LongTermDebtCurrent",
                                             "LongTermDebtAndCapitalLeaseObligationsCurrent"],
                                     fiscal_year)
        # Only add current if the total concept name suggests it's non-current only
        # Use heuristic: if current > 10% of total, it's likely already included
        if current_raw is not None and current_raw < total_debt_raw * 0.5:
            # Assume current not included; most single-concept filers include both
            # but LongTermDebt (noncurrent) filers need current added
            noncurrent_only = value_for_year(facts, ["LongTermDebtNoncurrent"], fiscal_year)
            if noncurrent_only is not None:
                total_debt_raw = noncurrent_only + current_raw
        total_debt = to_thousands(total_debt_raw)
    else:
        # Build from noncurrent + current components
        lt_raw = value_for_year(facts, ["LongTermDebtNoncurrent"], fiscal_year)
        st_raw = value_for_year(facts, ["ShortTermBorrowings", "DebtCurrent"], fiscal_year)
        if lt_raw is not None and st_raw is not None:
            total_debt = to_thousands(lt_raw + st_raw)
        elif lt_raw is not None:
            total_debt = to_thousands(lt_raw)
        elif st_raw is not None:
            total_debt = to_thousands(st_raw)
        else:
            total_debt = None

    # Equity: direct concept, or Assets - Liabilities
    equity = v([
        "StockholdersEquity",
        "StockholdersEquityIncludingPortionAttributableToNoncontrollingInterest",
    ])
    if equity is None and total_assets is not None:
        liabilities_raw = value_for_year(facts, ["Liabilities"], fiscal_year)
        if liabilities_raw is not None:
            # total_assets and liabilities_raw are both in whole dollars here
            assets_raw = value_for_year(facts, ["Assets"], fiscal_year)
            if assets_raw is not None:
                equity = to_thousands(assets_raw - liabilities_raw)

    # EBITDA from pre-built year-matched series
    ebitda_raw = ebitda_series.get(fiscal_year)
    ebitda = to_thousands(ebitda_raw)

    result: dict[str, Optional[float]] = {
        "revenue":         revenue,
        "ebitda":          ebitda,
        "net_income":      net_income,
        "total_assets":    total_assets,
        "total_debt":      total_debt,
        "cash":            cash,
        "equity":          equity,
        "pv10":            None,
        "proved_reserves": None,
    }

    if is_oil_gas:
        # PV-10: try XBRL concepts in order of specificity
        pv10_xbrl_concepts = [
            "StandardizedMeasureOfDiscountedFutureNetCashFlows",
            "StandardizedMeasureOfDiscountedFutureNetCashFlowsRelatingToProvedOilAndGasReserves",
            "DiscountedFutureNetCashFlowsRelatingToProvedOilAndGasReservesStandardizedMeasure",
            "FutureNetCashFlowsRelatingToProvedOilAndGasReservesTenPercentAnnualDiscountForEstimatedTimingOfCashFlows",
            "DiscountedFutureNetCashFlowsRelatingToProvedOilAndGasReserves10PercentAnnualDiscountForEstimatedTimingOfCashFlows",
        ]
        pv10_raw = value_for_year(facts, pv10_xbrl_concepts, fiscal_year)

        # Proved reserves: XBRL (energy units, usually MMBoe or Bcfe)
        reserves_raw = value_for_year(facts, [
            "ProvedDevelopedAndUndevelopedReserveNetEnergy",
            "ProvedDevelopedReservesBOE1",
        ], fiscal_year)

        # HTML scrape only when explicitly requested (most recent year, caller decides)
        if scrape_html and (pv10_raw is None or reserves_raw is None):
            accs = [accession]
            if original_accession != accession:
                accs.append(original_accession)
            for acc in accs:
                html_text = fetch_10k_html(cik, acc)
                if not html_text:
                    continue
                scale = detect_html_scale(html_text)
                pv10_html, res_html = scrape_pv10_and_reserves(html_text)
                if pv10_raw is None and pv10_html is not None:
                    pv10_raw = pv10_html * scale   # convert to whole dollars
                if reserves_raw is None:
                    reserves_raw = res_html
                # Stop if we found what we needed; otherwise try next accession
                if pv10_raw is not None and reserves_raw is not None:
                    break

        result["pv10"]            = to_thousands(pv10_raw)
        result["proved_reserves"] = reserves_raw    # NOT in $thousands — raw MMBoe/Bcfe

    return result


# ---------------------------------------------------------------------------
# Formatting helpers
# ---------------------------------------------------------------------------

def fmt_thousands(val: Optional[float], is_reserves: bool = False) -> str:
    """Format a value stored in $thousands for display."""
    if val is None:
        return "N/A"
    if is_reserves:
        return f"{val:,.1f}"
    # Display in $M (divide thousands by 1,000)
    m = val / 1_000
    return f"${m:,.1f}M"


def fmt_excel(val: Optional[float]) -> object:
    """Raw value for Excel (already in $thousands, or raw for reserves)."""
    return val if val is not None else ""


# ---------------------------------------------------------------------------
# Console output
# ---------------------------------------------------------------------------

def print_company_table(
    ticker: str,
    company_name: str,
    is_oil_gas: bool,
    years_data: list[tuple[str, dict]],
    is_trust: bool = False,
) -> None:
    year_labels = [f"FY{fy}" for fy, _ in years_data]
    labels = TRUST_METRIC_LABELS if is_trust else METRIC_LABELS
    show_og = is_oil_gas or is_trust  # trusts may have PV-10/reserves
    metrics = [k for k in METRIC_KEYS
               if show_og or k not in ("pv10", "proved_reserves")]
    # Trusts: hide EBITDA (always N/A) and Total Debt (always 0, not interesting)
    if is_trust:
        metrics = [k for k in metrics if k not in ("ebitda",)]

    header = ["Metric (USD Millions)"] + year_labels
    rows = []
    for key in metrics:
        label = labels[key]
        is_res = (key == "proved_reserves")
        row = [label] + [fmt_thousands(data.get(key), is_reserves=is_res)
                         for _, data in years_data]
        rows.append(row)

    width = 57 + 14 * len(years_data)
    print(f"\n{'=' * width}")
    trust_tag = " [Royalty Trust]" if is_trust else ""
    print(f"  {ticker.upper()} — {company_name}{trust_tag}")
    print(f"{'=' * width}")
    print(tabulate(rows, headers=header, tablefmt="rounded_outline"))
    if is_trust:
        print("\n  * Modified cash basis — no EBITDA; Equity = Trust Corpus end-of-year")
    else:
        print("\n  * EBITDA = Operating Income + D&A (same fiscal year)")
    if is_oil_gas or is_trust:
        print("  * PV-10 = Standardized Measure of Discounted Future Net Cash Flows")
        print("  * Proved Reserves in filing units (MMBoe or Bcfe)")
    print()


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def export_excel(
    all_results: list[tuple[str, str, bool, list[tuple[str, dict]], bool]],
    path: str,
) -> None:
    """
    Write one sheet per ticker. Each sheet has metrics as rows, fiscal years as columns.
    All dollar values are in $thousands; the sheet header row says so.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl not installed — run: pip3 install openpyxl", file=sys.stderr)
        return

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    # Colour palette
    HEADER_BG   = "1F3864"   # dark navy
    HEADER_FG   = "FFFFFF"
    SUBHDR_BG   = "2E75B6"   # medium blue
    SUBHDR_FG   = "FFFFFF"
    ALT_BG      = "EAF0FB"   # light blue alternating row
    OG_LABEL_BG = "FFF2CC"   # light yellow for oil/gas-only rows
    BORDER_CLR  = "BDD7EE"

    thin = Side(style="thin", color=BORDER_CLR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_cell(cell, bold=False, bg=None, fg="000000", align="left",
                   num_fmt=None, wrap=False):
        cell.font      = Font(bold=bold, color=fg, name="Calibri", size=10)
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        cell.border    = border
        if bg:
            cell.fill  = PatternFill("solid", fgColor=bg)
        if num_fmt:
            cell.number_format = num_fmt

    for ticker, company_name, is_oil_gas, years_data, is_trust in all_results:
        ws = wb.create_sheet(title=ticker.upper()[:31])
        ws.sheet_view.showGridLines = False

        fiscal_years = [fy for fy, _ in years_data]
        labels = TRUST_METRIC_LABELS if is_trust else METRIC_LABELS
        show_og = is_oil_gas or is_trust
        metrics = [k for k in METRIC_KEYS
                   if show_og or k not in ("pv10", "proved_reserves")]
        if is_trust:
            metrics = [k for k in metrics if k not in ("ebitda",)]

        # --- Row 1: company banner ---
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=1 + len(fiscal_years))
        trust_tag = " [Royalty Trust]" if is_trust else ""
        banner = ws.cell(1, 1, f"{ticker.upper()} — {company_name}{trust_tag}")
        style_cell(banner, bold=True, bg=HEADER_BG, fg=HEADER_FG, align="center")
        ws.row_dimensions[1].height = 20

        # --- Row 2: column headers (Metric | FY20XX ...) ---
        ws.cell(2, 1, "Metric")
        style_cell(ws.cell(2, 1), bold=True, bg=SUBHDR_BG, fg=SUBHDR_FG, align="center")
        for col_i, fy in enumerate(fiscal_years, start=2):
            c = ws.cell(2, col_i, f"FY{fy}")
            style_cell(c, bold=True, bg=SUBHDR_BG, fg=SUBHDR_FG, align="center")
        ws.row_dimensions[2].height = 16

        # --- Row 3: unit note ---
        ws.merge_cells(start_row=3, start_column=1,
                       end_row=3, end_column=1 + len(fiscal_years))
        unit_note = ws.cell(3, 1, "All dollar values in USD thousands ($000s), except Proved Reserves (MMBoe/Bcfe)")
        style_cell(unit_note, bg="F2F2F2", fg="595959", align="center", wrap=True)
        ws.row_dimensions[3].height = 14

        # --- Data rows ---
        for row_i, key in enumerate(metrics, start=4):
            is_res     = (key == "proved_reserves")
            is_og_only = key in ("pv10", "proved_reserves")
            label_bg   = OG_LABEL_BG if is_og_only else (ALT_BG if row_i % 2 == 0 else None)

            label_cell = ws.cell(row_i, 1, labels[key])
            style_cell(label_cell, bold=True, bg=label_bg, align="left")

            for col_i, (fy, data) in enumerate(years_data, start=2):
                raw = data.get(key)
                cell = ws.cell(row_i, col_i)

                if raw is None:
                    cell.value = None
                elif is_res:
                    cell.value = raw           # MMBoe, no dollar format
                    cell.number_format = '#,##0.0'
                else:
                    cell.value = raw           # already in $thousands
                    cell.number_format = '#,##0'

                style_cell(cell, bg=label_bg, align="right")

            ws.row_dimensions[row_i].height = 15

        # --- Column widths ---
        ws.column_dimensions[get_column_letter(1)].width = 26
        for col_i in range(2, 2 + len(fiscal_years)):
            ws.column_dimensions[get_column_letter(col_i)].width = 18

        # --- Notes section ---
        note_row = 4 + len(metrics) + 1
        ws.merge_cells(start_row=note_row, start_column=1,
                       end_row=note_row, end_column=1 + len(fiscal_years))
        if is_trust:
            notes = [
                "Notes:",
                "• Modified cash basis — EBITDA not applicable; Equity = Trust Corpus end-of-year.",
                "• Source: SEC EDGAR 10-K HTML (trusts do not file XBRL); multi-year from comparative statements.",
                "• PV-10 = Standardized Measure of Discounted Future Net Cash Flows (most recent year only).",
                "• Proved Reserves in filing units (MMBoe or Bcfe).",
            ]
        else:
            notes = [
                "Notes:",
                "• EBITDA estimated as Operating Income + D&A, matched within the same fiscal year.",
                "• Source: SEC EDGAR XBRL data (data.sec.gov); oil & gas supplemental via HTML parsing.",
            ]
            if is_oil_gas:
                notes.append("• PV-10 = Standardized Measure of Discounted Future Net Cash Flows.")
                notes.append("• Proved Reserves in filing units (MMBoe or Bcfe — verify per filing).")
        note_cell = ws.cell(note_row, 1, "  ".join(notes))
        style_cell(note_cell, fg="595959", bg="F9F9F9", wrap=True)
        ws.row_dimensions[note_row].height = 60

    wb.save(path)
    print(f"Excel file saved: {path}")


# ---------------------------------------------------------------------------
# Main per-ticker extraction
# ---------------------------------------------------------------------------

def extract_ticker(
    ticker: str, num_years: int
) -> tuple[str, str, bool, list[tuple[str, dict]], bool]:
    """
    Fetch and extract data for one ticker.
    Returns (ticker, company_name, is_oil_gas, [(fiscal_year, metrics), ...], is_trust).
    """
    print(f"\n{'─'*55}")
    print(f"  {ticker.upper()}")
    print(f"{'─'*55}")

    cik, company_name = get_cik(ticker)
    print(f"  Company : {company_name}")
    print(f"  CIK     : {int(cik)}")

    info       = get_company_info(cik)
    sic        = str(info.get("sic", ""))
    is_oil_gas = sic in OIL_GAS_SIC
    if is_oil_gas:
        print(f"  SIC {sic} — Oil & Gas company detected.")

    filings = get_10k_filings(info, n=num_years)
    print(f"  Found {len(filings)} filing(s): " +
          ", ".join(f"FY{f['fiscal_year']} ({f['date']})" for f in filings))

    print("  Fetching XBRL facts...")
    facts = get_company_facts(cik)

    # --- Trust detection ---
    trust_flag = is_royalty_trust(company_name, sic, facts)
    if trust_flag:
        print("  Royalty Trust detected — switching to HTML scrape mode.")
        return _extract_trust(ticker, company_name, cik, is_oil_gas, filings, num_years)

    # --- Standard XBRL path ---
    ebitda_series = _build_ebitda_series(facts)
    years_data: list[tuple[str, dict]] = []
    for i, filing in enumerate(filings):
        fy   = filing["fiscal_year"]
        acc  = filing["accession"]
        orig = filing["original_accession"]
        scrape_html = is_oil_gas and (i == 0)
        if scrape_html:
            print(f"  Fetching 10-K HTML for FY{fy} (PV-10 / reserves)...")

        metrics = extract_year(
            facts              = facts,
            fiscal_year        = fy,
            ebitda_series      = ebitda_series,
            is_oil_gas         = is_oil_gas,
            cik                = cik,
            accession          = acc,
            original_accession = orig,
            scrape_html        = scrape_html,
        )
        years_data.append((fy, metrics))

    return ticker, company_name, is_oil_gas, years_data, False


def _extract_trust(
    ticker: str,
    company_name: str,
    cik: str,
    is_oil_gas: bool,
    filings: list[dict],
    num_years: int,
) -> tuple[str, str, bool, list[tuple[str, dict]], bool]:
    """
    Trust-mode extraction: scrape HTML from each 10-K filing.
    The most recent 10-K often contains 3-year comparative income data
    and 2-year comparative balance sheet data in a single document.
    We fetch up to two filings to cover three full years.
    """
    all_year_data: dict[str, dict] = {}

    for i, filing in enumerate(filings):
        fy  = filing["fiscal_year"]
        acc = filing["accession"]
        orig = filing["original_accession"]
        print(f"  Fetching 10-K HTML for FY{fy}...")

        accs = [acc]
        if orig != acc:
            accs.append(orig)

        html_text = ""
        for a in accs:
            html_text = fetch_10k_html(cik, a)
            if html_text:
                break

        if not html_text:
            print(f"    [warn] Could not fetch HTML for FY{fy}.")
            all_year_data.setdefault(fy, _empty_metrics())
            continue

        scraped = scrape_trust_all_years(html_text, fy, is_oil_gas)
        for yr, m in scraped.items():
            # Don't overwrite data already found from a more-recent filing
            if yr not in all_year_data:
                all_year_data[yr] = m

        # The most recent filing already gives us 3 years of income data;
        # stop fetching HTML unless we still have gaps
        wanted = {f["fiscal_year"] for f in filings}
        if wanted.issubset(all_year_data.keys()):
            break

    # Build ordered list matching the filings order
    years_data = [(f["fiscal_year"], all_year_data.get(f["fiscal_year"], _empty_metrics()))
                  for f in filings]

    return ticker, company_name, is_oil_gas, years_data, True


def _empty_metrics() -> dict:
    return {k: None for k in METRIC_KEYS}


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Extract key financials from SEC EDGAR 10-K filings.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "Examples:\n"
            "  python3 sec_10k_extractor.py AAPL\n"
            "  python3 sec_10k_extractor.py AAPL,MSFT,GOOG --years 3\n"
            "  python3 sec_10k_extractor.py XOM,CVX --export oilgas.xlsx\n"
        )
    )
    parser.add_argument(
        "tickers",
        help="One or more ticker symbols, comma-separated (e.g. AAPL,MSFT,XOM)",
    )
    parser.add_argument(
        "--years", "-y",
        type=int, default=3,
        help="Number of fiscal years to extract (default: 3)",
    )
    parser.add_argument(
        "--export", "-e",
        metavar="FILE.xlsx",
        help="Export results to an Excel file",
    )
    args = parser.parse_args()

    tickers = [t.strip().upper() for t in args.tickers.split(",") if t.strip()]
    all_results = []
    errors = []

    for ticker in tickers:
        try:
            result = extract_ticker(ticker, args.years)
            all_results.append(result)
            t, company_name, is_oil_gas, years_data, is_trust = result
            print_company_table(t, company_name, is_oil_gas, years_data, is_trust)
        except Exception as exc:
            errors.append((ticker, str(exc)))
            print(f"\n  [ERROR] {ticker}: {exc}", file=sys.stderr)

    if args.export and all_results:
        export_excel(all_results, args.export)

    if errors:
        print("\nFailed tickers:", ", ".join(t for t, _ in errors), file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
